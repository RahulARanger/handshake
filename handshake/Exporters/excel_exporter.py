from typing import Optional, Any, Union, Tuple
from handshake.services.SchedularService.constants import exportExportFileName, JobType
from handshake.services.SchedularService.register import warn_about_test_run
from handshake.services.DBService.models.dynamic_base import TaskBase
from handshake.services.DBService.models.result_base import (
    SuiteBase,
    SuiteType,
)
from handshake.services.DBService.models.attachmentBase import AssertBase
from handshake.services.DBService.models.static_base import StaticBase
from handshake.Exporters.exporter import Exporter
from handshake.services.DBService.lifecycle import attachment_folder
from loguru import logger
from pathlib import Path
from aiofiles.os import mkdir
from datetime import datetime
from tortoise.expressions import Q

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell import Cell
    from openpyxl.styles.alignment import Alignment
    from openpyxl.comments.comments import Comment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.formatting import ConditionalFormatting
    from openpyxl.styles.fills import (
        FILL_SOLID,
        PatternFill,
        FILL_PATTERN_DARKHORIZONTAL,
    )

    excel_export = True
except ImportError:
    Workbook = Any
    Worksheet = Any
    ConditionalFormatting = Any
    excel_export = False
    Cell = Any

# PLEASE DO NOT TYPECAST EXCEL RELATED THINGS UNLESS WE HANDLE IT IN EXCEPTION


def save_datetime_in_excel(obj: datetime, cell: Optional[Cell] = None):
    full_text = obj.strftime("%Y-%m-%d %I:%M:%S %p")
    if not cell:
        return full_text
    cell.comment = Comment(full_text, "Handshake", width=150, height=40)
    return obj.strftime("%I:%M:%S %p")


def calc_percentage(num, deno):
    return (num / deno) if deno else 0


class ExcelExporter(Exporter):
    template: Workbook
    standing_format: ConditionalFormatting
    percentage_format: ConditionalFormatting
    parent_links = {}
    test_link = {}

    test_summary_row = 3
    test_summary_col = 4
    test_summary_files_row = -1

    def __init__(self, db_path: Path, dev_run: bool = False):
        super().__init__(dev_run)
        self.save_in = attachment_folder(db_path)

    def prepare(self):
        self.template = load_workbook(Path(__file__).parent / "Template.xlsx")

    async def start_exporting(
        self,
        run_id: Optional[str] = None,
        skip_project_summary: bool = False,
        skip_prepare: bool = False,
    ):
        if not excel_export:
            return logger.error(
                "Please install excel exporter using"
                " pip install handshake[excel-export] or poetry add handshake[excel-export]"
            )
        task = await TaskBase.filter(type=JobType.EXPORT_EXCEL, test_id=run_id).first()
        try:
            await super().start_exporting(run_id, True)
        except Exception as error:
            logger.exception("Failed to export this test run in excel due to {}", error)
            await warn_about_test_run(
                run_id,
                "Failed to export this test run in excel",
                task.type,
                error=repr(error),
                task=task.ticketID,
            )
            task.processed = True
            await task.save()
            return None

        self.save_in.mkdir(exist_ok=True)
        if not run_id:
            return None

        not (self.save_in / str(run_id)).exists() and await mkdir(
            self.save_in / str(run_id)
        )
        self.template.save(self.save_in / str(run_id) / exportExportFileName)

        task.processed = True
        task.picked = False
        await task.save()
        return None

    def completed(self):
        logger.info(
            "Excel is available in Attachments folder, Please check at the {}",
            self.save_in,
        )

    async def export_test_run_summary(self, test_id: str, summary):
        index_sheet = self.template["Index"]

        edit_cell(
            index_sheet,
            self.test_summary_row,
            self.test_summary_col,
            summary["projectName"],
            True,
        )
        self.test_summary_row += 1

        standing_cell = edit_cell(
            index_sheet,
            self.test_summary_row,
            self.test_summary_col,
            summary["standing"].lower().capitalize(),
            True,
        )
        self.test_summary_row += 1

        self.standing_format, self.percentage_format, *_ = list(
            index_sheet.conditional_formatting
        )

        standing_cell.comment = Comment(
            f"Run Status: {summary['status'].lower().capitalize()}\n"
            f"Exit Code: {summary['exitCode']}",
            summary["framework"],
            width=200,
        )

        edit_cell(
            index_sheet,
            self.test_summary_row,
            self.test_summary_col,
            format_duration(summary["duration"] / 1000),
        )
        self.test_summary_row += 1

        edit_cell(
            index_sheet,
            self.test_summary_row,
            self.test_summary_col,
            summary["suites"],
        )
        self.test_summary_row += 1

        edit_cell(
            index_sheet,
            self.test_summary_row,
            self.test_summary_col,
            summary["tests"],
        )
        self.test_summary_row += 1

        for status in ("passed", "failed", "skipped", "xfailed", "xpassed"):
            edit_cell(
                index_sheet,
                self.test_summary_row,
                self.test_summary_col,
                summary[status],
            )
            self.test_summary_row += 1  # at the end of the loop for overall percentage

        edit_cell(
            index_sheet,
            self.test_summary_row,
            self.test_summary_col,
            calc_percentage(
                summary["passedSuites"] + summary["skippedSuites"],
                summary["suites"],
            ),
        )
        self.test_summary_row += 2  # files, platform
        self.test_summary_files_row = self.test_summary_row - 1

        edit_cell(
            index_sheet,
            self.test_summary_row,
            self.test_summary_col,
            summary["platform"],
            True,
        )
        self.test_summary_row += 1

        edit_cell(
            index_sheet,
            self.test_summary_row,
            self.test_summary_col,
            summary["framework"],
        )
        # starting with the timeline table
        self.test_summary_row += 3
        edit_cell(
            index_sheet,
            self.test_summary_row,
            self.test_summary_col,
            save_datetime_in_excel(datetime.fromisoformat(summary["started"])),
        )
        self.test_summary_row += 1

        edit_cell(
            index_sheet,
            self.test_summary_row,
            self.test_summary_col,
            save_datetime_in_excel(datetime.fromisoformat(summary["ended"])),
        )
        self.test_summary_row += 1

        edit_cell(
            index_sheet,
            self.test_summary_row,
            self.test_summary_col,
            save_datetime_in_excel(datetime.now()),
            resize=True,
        )
        self.test_summary_row += 1

    async def export_run_page(self, run_id: str, skip_recent_suites: bool = False):
        await super().export_run_page(run_id, True)

    async def export_project_summary(self, run_feed, projects_feed): ...

    async def export_overview_of_test_run(self, run_id: str, summary):
        index_sheet = self.template["Index"]

        edit_cell(
            index_sheet,
            self.test_summary_files_row,
            self.test_summary_col,
            summary["aggregated"]["files"],
        )

    async def export_all_suites(self, run_id: str, export_suite_wise: bool = True):
        await super().export_all_suites(run_id, False)

    async def export_all_suites_of_test_run(self, run_id, all_suites):
        if not all_suites:
            return
        suites_sheet = self.template["Test Scenarios"]
        table_style = TableStyleInfo(name="TableStyleMedium23", showRowStripes=True)

        columns = (
            "title",
            "aliasID",
            "total_duration",
            "description",
            "standing",
            "tests",
            "rollup_passed",
            "rollup_failed",
            "rollup_skipped",
            "rollup_xfailed",
            "rollup_xpassed",
            "numberOfErrors",
            "error",
            "started",
            "ended",
            "parent",
            "simplified",
            "file",
            "duration",
            "setup_duration",
            "teardown_duration",
            "retried_later",
            "Parent Alias",
        )
        alias = {
            "aliasID": "Alias",
            "numberOfErrors": "Errors",
            "duration": "test_duration",
            "tests": "passed%",
            "simplified": "Ran with",
            "standing": "status",
            "rollup_passed": "Passed",
            "rollup_failed": "Failed",
            "rollup_skipped": "Skipped",
            "rollup_xfailed": "XFailed",
            "rollup_xpassed": "XPassed",
            "retried_later": "Retried Later",
            "setup_duration": "Setup",
            "teardown_duration": "Teardown",
        }

        suites_sheet.freeze_panes = "B1"  # freeze first col (title)

        for row_index, suite in enumerate(all_suites):
            for header_index, header in enumerate(columns):
                if not row_index:
                    suites_sheet.cell(1, header_index + 1).value = alias.get(
                        header, header
                    ).capitalize()

                cell = suites_sheet.cell(row_index + 2, header_index + 1)
                cell.alignment = Alignment("right")
                match header:
                    case "Parent Alias":
                        value: Union[Any, Tuple[str, str, str], bool] = (
                            self.parent_links.get(suite["parent"], False)
                        )
                    case _:
                        value = suite[header]

                to_save = value
                resize = True
                match header:
                    case "started":
                        to_save = save_datetime_in_excel(
                            datetime.fromisoformat(value), cell
                        )
                    case "ended":
                        to_save = save_datetime_in_excel(
                            datetime.fromisoformat(value), cell
                        )
                    case "duration":
                        to_save = format_duration(value / 1000)
                        resize = (
                            False  # else it considers the length of the formula string
                        )
                        resize_col(suites_sheet, cell, 10)
                    case "total_duration":
                        to_save = format_duration(value / 1000)
                        resize = False
                        resize_col(suites_sheet, cell, 6)
                    case "setup_duration":
                        to_save = format_duration(value / 1000)
                        resize = False
                        resize_col(suites_sheet, cell, 6)
                    case "teardown_duration":
                        to_save = format_duration(value / 1000)
                        resize = False
                        resize_col(suites_sheet, cell, 7)
                    case "tests":
                        to_save = calc_percentage(
                            suite["rollup_passed"] + suite["rollup_skipped"],
                            suite["rollup_tests"],
                        )
                        cell.number_format = "0%"
                        copy_format_to_cell(suites_sheet, cell, self.percentage_format)
                        resize = False
                    case "title":
                        resize = True
                        self.parent_links[suite["suiteID"]] = (
                            f"#'{suites_sheet.title}'!{cell.column_letter}{cell.row}",
                            value,
                            suite["aliasID"],
                        )
                    case "standing":
                        to_save = value.lower().capitalize()
                        resize = True
                        copy_format_to_cell(suites_sheet, cell, self.standing_format)
                    case "title":
                        cell.alignment = Alignment(wrapText=True)
                        resize_col(suites_sheet, cell, 30)
                        resize = False
                    case "description":
                        if value:
                            cell.alignment = Alignment(wrapText=True)
                            resize_col(suites_sheet, cell, 30)
                            resize = False
                    case "error":
                        if value:
                            resize_col(suites_sheet, cell, 30)
                            resize = False
                    case "parent":
                        to_save = "〰️"
                        cell.alignment = Alignment("center", "center")
                        if value and self.parent_links[value]:
                            to_save = f"{suite['parent_title']} 🔗"
                            cell.hyperlink = self.parent_links[value][0]
                            resize = True
                            add_comment_to_cell(cell, self.parent_links[value][1])
                    case "Parent Alias":
                        to_save = "〰️"
                        cell.alignment = Alignment("center", "center")
                        if value:
                            to_save = f"{value[-1]} 🔗"
                            cell.hyperlink = value[0]
                            resize = True
                            add_comment_to_cell(cell, value[-1])
                    case "retried_later":
                        to_save = "YES" if value else "NO"
                    case "file":
                        add_comment_to_cell(cell, value)
                        resize = False
                        resize_col(suites_sheet, cell, 10)
                    case _:
                        to_save = value
                        resize = False
                        resize_col(suites_sheet, cell, 6)

                edit_cell(
                    suites_sheet, row_index + 2, header_index + 1, to_save, resize
                )

        suites_sheet.add_table(
            Table(
                ref=f"A1:{suites_sheet.cell(1, 1).offset(0, len(columns) - 1).column_letter}"
                f"{len(all_suites) + 1}",
                displayName="Test_Scenarios",
                tableStyleInfo=table_style,
            )
        )

    @staticmethod
    def export_test_query(based_on: str):
        return (
            SuiteBase.filter(
                ~Q(suiteType=SuiteType.SUITE) & Q(session__test_id=based_on)
            ),
            AssertBase.filter(entity__session__test_id=based_on),
            StaticBase.filter(entity__session__test_id=based_on),
        )

    async def export_tests(self, run_id, suite_id, tests):
        if not tests:
            return
        test_sheet = self.template["Test Cases"]
        hooks_sheet = self.template["Hooks"]
        table_style = TableStyleInfo(name="TableStyleMedium23", showRowStripes=True)

        columns = (
            "title",
            "aliasID",
            "total_duration",
            "suiteType",
            "description",
            "standing",
            "assertions",
            "numberOfErrors",
            "error",
            "started",
            "ended",
            "parent",
            "file",
            "duration",
            "setup_duration",
            "teardown_duration",
            "retried_later",
            "Parent Alias",
        )
        alias = {
            "aliasID": "alias",
            "duration": "test_duration",
            "numberOfErrors": "Errors",
            "simplified": "Ran with",
            "standing": "status",
            "retried_later": "Retried Later",
            "setup_duration": "Setup",
            "teardown_duration": "Teardown",
            "suiteType": "Type",
        }

        test_sheet.freeze_panes = "B1"  # freeze first col (title)
        hooks_sheet.freeze_panes = "B1"  # freeze first col (title)

        test_rows = 2
        hooks_row = 2

        for row_index, test in enumerate(tests):
            is_test = test["suiteType"] == SuiteType.TEST

            for header_index, header in enumerate(columns):
                if not row_index:
                    header_value = alias.get(header, header).capitalize()
                    test_sheet.cell(1, header_index + 1).value = header_value
                    hooks_sheet.cell(1, header_index + 1).value = header_value

                sheet = test_sheet if is_test else hooks_sheet
                cell = sheet.cell(test_rows if is_test else hooks_row, header_index + 1)
                cell.alignment = Alignment("right")

                match header:
                    case "Parent Alias":
                        value: Union[Any, Tuple[str, str, str], bool] = (
                            self.parent_links.get(test["parent"], False)
                        )
                    case _:
                        value = test[header]

                resize = True
                to_save = value

                match header:
                    case "started":
                        to_save = save_datetime_in_excel(
                            datetime.fromisoformat(value), cell
                        )
                    case "ended":
                        to_save = save_datetime_in_excel(
                            datetime.fromisoformat(value), cell
                        )
                    case "duration":
                        to_save = format_duration(value / 1000)
                        resize = (
                            False  # else it considers the length of the formula string
                        )
                        resize_col(sheet, cell, 10)
                    case "total_duration":
                        to_save = format_duration(value / 1000)
                        resize = False
                        resize_col(sheet, cell, 6)
                    case "setup_duration":
                        to_save = format_duration(value / 1000)
                        resize = False
                        resize_col(sheet, cell, 6)
                    case "teardown_duration":
                        to_save = format_duration(value / 1000)
                        resize = False
                        resize_col(sheet, cell, 7)
                    case "title":
                        resize = True
                        self.parent_links[test["suiteID"]] = (
                            f"#'{test_sheet.title}'!{cell.column_letter}{cell.row}",
                            value,
                            test["aliasID"],
                        )
                    case "standing":
                        to_save = value.lower().capitalize()
                        resize = True
                        copy_format_to_cell(sheet, cell, self.standing_format)
                    case "title":
                        cell.alignment = Alignment(wrapText=True)
                        resize_col(sheet, cell, 30)
                        resize = False
                    case "description":
                        if value:
                            cell.alignment = Alignment(wrapText=True)
                            resize_col(sheet, cell, 30)
                            resize = False
                    case "error":
                        if value:
                            resize_col(sheet, cell, 30)
                            resize = False
                    case "parent":
                        to_save = ""
                        index = (
                            test_rows
                            if test["suiteType"] == SuiteType.SETUP
                            else (test_rows - 1)
                        )
                        link = self.parent_links.get(
                            value,
                            (f"#'{test_sheet.title}'!A{index}", "", ""),
                        )
                        if value and link:
                            to_save = f"{test['parent_title']} 🔗"
                            resize = True
                            cell.hyperlink = link[0]
                            cell.alignment = Alignment("center", "center")

                            if link[1]:
                                add_comment_to_cell(cell, link[1])
                    case "Parent Alias":
                        to_save = ""
                        index = (
                            test_rows
                            if test["suiteType"] == SuiteType.SETUP
                            else (test_rows - 1)
                        )
                        link = (
                            value
                            if value
                            else (f"#'{test_sheet.title}'!A{index}", "", "")
                        )
                        if value and link:
                            to_save = f"{link[2]} 🔗"
                            resize = True
                            cell.hyperlink = link[0]
                            cell.alignment = Alignment("center", "center")

                            if link[2]:
                                add_comment_to_cell(cell, link[2])
                    case "retried_later":
                        to_save = "YES" if value else "NO"
                    case "file":
                        add_comment_to_cell(cell, value)
                        resize = False
                        resize_col(sheet, cell, 10)
                    case _:
                        to_save = value
                        resize = False
                        resize_col(sheet, cell, 6)

                edit_cell(
                    sheet,
                    cell.row,
                    cell.column,
                    to_save,
                    resize,
                )
            if is_test:
                test_rows += 1
            else:
                hooks_row += 1

        test_rows > 2 and test_sheet.add_table(
            Table(
                ref=f"A1:{test_sheet.cell(1, 1).offset(0, len(columns) - 1).column_letter}"
                f"{test_rows - 1}",
                displayName="Test_Cases",
                tableStyleInfo=table_style,
            )
        )
        hooks_row > 2 and hooks_sheet.add_table(
            Table(
                ref=f"A1:{hooks_sheet.cell(1, 1).offset(0, len(columns) - 1).column_letter}"
                f"{hooks_row - 1}",
                displayName="Hooks",
                tableStyleInfo=table_style,
            )
        )

    async def export_retries_map(self, run_id, suite_id, retried_map): ...

    async def export_attachments(
        self, run_id, suite_id, assertion_records, written_records
    ):
        if not assertion_records:
            return
        assertion_sheet = self.template["Assertions"]
        table_style = TableStyleInfo(name="TableStyleMedium23", showRowStripes=True)

        columns = (
            "title",
            "raw",
            "entity_id",
            "passed",
            "interval",
            "wait",
            "Test Alias",
        )
        alias = {"entity_id": "entity", "raw": "description"}

        assertion_sheet.freeze_panes = "B1"  # freeze first col (title)
        fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type=FILL_SOLID
        )
        assertions = [_ for __ in assertion_records for _ in assertion_records[__]]

        for row_index, assertion in enumerate(assertions):
            for header_index, header in enumerate(columns):
                if not row_index:
                    assertion_sheet.cell(1, header_index + 1).value = alias.get(
                        header, header
                    ).capitalize()

                cell = assertion_sheet.cell(row_index + 2, header_index + 1)
                cell.alignment = Alignment("right")
                if not assertion["passed"]:
                    cell.fill = fill

                match header:
                    case "Test Alias":
                        value = self.parent_links.get(assertion["entity_id"], False)
                    case _:
                        value = assertion[header]
                to_save = value
                resize = True
                match header:
                    case "title":
                        cell.alignment = Alignment(wrapText=True)
                        resize_col(assertion_sheet, cell, 30)
                        resize = False
                    case "raw":
                        if value:
                            cell.alignment = Alignment(wrapText=True)
                            resize_col(assertion_sheet, cell, 30)
                            resize = False
                    case "entity_id":
                        to_save = "〰️"
                        cell.alignment = Alignment("center", "center")
                        link, tip, parent_alias = self.parent_links.get(
                            value, ("", "", "")
                        )
                        if value and link:
                            to_save = f"{tip} 🔗"
                            cell.hyperlink = link
                            resize = True
                            add_comment_to_cell(cell, tip)
                    case "Test Alias":
                        to_save = "〰️"
                        cell.alignment = Alignment("center", "center")
                        link, tip, parent_alias = value
                        if value and link:
                            to_save = f"{parent_alias} 🔗"
                            cell.hyperlink = link
                            resize = True
                            add_comment_to_cell(cell, parent_alias)
                    case "passed":
                        to_save = "YES" if value else "NO"
                    case _:
                        to_save = value
                        resize = False
                        resize_col(assertion_sheet, cell, 6)

                edit_cell(
                    assertion_sheet, row_index + 2, header_index + 1, to_save, resize
                )

        assertion_sheet.add_table(
            Table(
                ref=f"A1:{assertion_sheet.cell(1, 1).offset(0, len(columns) - 1).column_letter}"
                f"{len(assertions) + 1}",
                displayName="Assertions",
                tableStyleInfo=table_style,
            )
        )


def edit_cell(
    sheet: Worksheet, row: int, col: int, value: str, resize: Optional[bool] = False
):
    cell = sheet.cell(row, col)
    cell.value = value
    if value and resize:
        resize_col(sheet, cell, len(value))
    return cell


def resize_col(sheet: Worksheet, cell: Cell, length: int):
    sheet.column_dimensions[cell.column_letter].width = max(
        sheet.column_dimensions[cell.column_letter].width, length
    )


def format_duration(value: int):
    return f"""=IF({value}>=3600, TEXT(INT({value}/3600), "0") & " hrs " & 
                TEXT(INT(MOD({value}, 3600)/60), "0") & " mins " & TEXT(MOD({value}, 60), "0") & " secs",
                IF({value}>=60, TEXT(INT({value}/60), "0") & " mins " & TEXT(MOD({value}, 60), "0") & " secs",
                IF({value}>=1, TEXT({value}, "0") & " secs", TEXT({value}*1000, "0") & " ms")))"""


def copy_format_to_cell(
    sheet: Worksheet, cell: Cell, formatting_rules: ConditionalFormatting
):
    for rule in formatting_rules.rules:
        sheet.conditional_formatting.add(
            f"{cell.column_letter}{cell.row}",
            rule,
        )


def add_comment_to_cell(cell: Cell, value: str):
    cell.comment = Comment(value, "Handshake")
    cell.comment.width = 150
    cell.comment.height = 30 + (len(value) / 30) * 30
